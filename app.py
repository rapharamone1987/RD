import streamlit as st
import pandas as pd
import plotly.express as px
import io
from datetime import datetime, timedelta
from fpdf import FPDF
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# 1. Configuração da página e Estado da Sessão (Padrão PT-BR)
st.set_page_config(page_title="Matriz GUT + RACI + Cronograma", layout="wide")
st.title("📊 Gestão Estratégica: GUT + RACI + Cronograma")
st.markdown("Priorize tarefas, atribua responsabilidades e gerencie os prazos em um único lugar.")

if "dados_estrategicos" not in st.session_state:
    hoje = datetime.today()
    st.session_state.dados_estrategicos = pd.DataFrame([
        {
            "Tarefa": "Realizar o inventário anual de bens móveis e verificação de plaquetas patrimoniais", 
            "G": 4, "U": 4, "T": 3,
            "Aprovador (A)": "Diretor de Administração", "Responsável (R)": "Equipe de Patrimônio",
            "Consultados (C)": "Contabilidade", "Informados (I)": "Auditoria Interna",
            "Início": hoje.strftime("%Y-%m-%d"), "Conclusão": (hoje + timedelta(days=15)).strftime("%Y-%m-%d")
        },
        {
            "Tarefa": "Corrigir inconsistências em relatórios de depreciação contábil do encerramento do exercício", 
            "G": 5, "U": 3, "T": 4,
            "Aprovador (A)": "Chefe de Finanças", "Responsável (R)": "Contador Líder",
            "Consultados (C)": "TI (Suporte)", "Informados (I)": "Diretoria Executiva",
            "Início": hoje.strftime("%Y-%m-%d"), "Conclusão": (hoje + timedelta(days=7)).strftime("%Y-%m-%d")
        }
    ])

# Força a limpeza absoluta de nulos
st.session_state.dados_estrategicos = st.session_state.dados_estrategicos.fillna("").astype(str)
st.session_state.dados_estrategicos["G"] = pd.to_numeric(st.session_state.dados_estrategicos["G"], errors='coerce').fillna(3).astype(int)
st.session_state.dados_estrategicos["U"] = pd.to_numeric(st.session_state.dados_estrategicos["U"], errors='coerce').fillna(3).astype(int)
st.session_state.dados_estrategicos["T"] = pd.to_numeric(st.session_state.dados_estrategicos["T"], errors='coerce').fillna(3).astype(int)

# 2. Formulário de Cadastro
with st.form("nova_tarefa_form", clear_on_submit=True):
    st.subheader("🆕 Cadastrar Nova Tarefa Estratégica")
    c1, c2, c3, c4 = st.columns([4, 2, 2, 2])
    with c1: descricao = st.text_input("Descrição da Tarefa:", placeholder="Digite a ação...")
    with c2: g = st.selectbox("Gravidade (G):", [1, 2, 3, 4, 5], index=2)
    with c3: u = st.selectbox("Urgência (U):", [1, 2, 3, 4, 5], index=2)
    with c4: t = st.selectbox("Tendência (T):", [1, 2, 3, 4, 5], index=2)
    
    st.markdown("**Atribuição de Papéis (Matriz RACI):**")
    r1, r2, r3, r4 = st.columns(4)
    with r1: aprovador = st.text_input("Aprovador (A):", placeholder="Ex: Diretor da Área").strip()
    with r2: responsavel = st.text_input("Responsável (R):", placeholder="Ex: Equipe X").strip()
    with r3: consultados = st.text_input("Consultados (C):", placeholder="Ex: Setor de TI").strip()
    with r4: informados = st.text_input("Informados (I):", placeholder="Ex: Secretário").strip()
        
    st.markdown("**Cronograma de Execução:**")
    dt1, dt2 = st.columns(2)
    with dt1: data_inicio = st.date_input("Data de Início:", datetime.today(), format="DD/MM/YYYY")
    with dt2: data_fim = st.date_input("Prazo de Conclusão:", datetime.today() + timedelta(days=10), format="DD/MM/YYYY")
    submit = st.form_submit_button("Adicionar à Matriz de Trabalho")

if submit and descricao:
    if data_inicio > data_fim:
        st.error("Erro: A data de início não pode ser posterior ao prazo de conclusão!")
    else:
        novo = pd.DataFrame([{"Tarefa": descricao, "G": int(g), "U": int(u), "T": int(t), "Aprovador (A)": aprovador, "Responsável (R)": responsavel, "Consultados (C)": consultados, "Informados (I)": informados, "Início": data_inicio.strftime("%Y-%m-%d"), "Conclusão": data_fim.strftime("%Y-%m-%d")}])
        st.session_state.dados_estrategicos = pd.concat([st.session_state.dados_estrategicos, novo], ignore_index=True).fillna("")
        st.rerun()

# 3. Processamento Geral dos Dados
df = st.session_state.dados_estrategicos.copy()
df["Score"] = df["G"] * df["U"] * df["T"]
df = df.sort_values(by="Score", ascending=False).reset_index(drop=True)
df_par_grafico = df.copy()

df["Início"] = pd.to_datetime(df["Início"]).dt.strftime("%d/%m/%Y")
df["Conclusão"] = pd.to_datetime(df["Conclusão"]).dt.strftime("%d/%m/%Y")
ordem = ["Tarefa", "G", "U", "T", "Score", "Início", "Conclusão", "Aprovador (A)", "Responsável (R)", "Consultados (C)", "Informados (I)"]
df = df[ordem]

# 4. Classe Customizada do PDF Executivo
class PDFExecutivo(FPDF):
    def header(self):
        self.set_draw_color(180, 180, 180)
        self.rect(10, 10, 277, 190, "D")
        self.set_fill_color(27, 94, 32) 
        self.rect(12, 12, 273, 14, "F")
        self.set_xy(15, 14)
        self.set_font("Arial", "B", 14)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, "PLANO DE AÇÃO E GOVERNANÇA ESTRATÉGICA", align="L")
        
    def footer(self):
        self.set_fill_color(44, 160, 90) 
        self.rect(12, 191, 273, 1.5, "F")
        self.set_y(193)
        self.set_x(15)
        self.set_font("Arial", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(100, 5, f"Documento Institucional - Emitido em: {datetime.today().strftime('%d/%m/%Y')}", align="L")
        self.set_x(240)
        self.cell(40, 5, f"Página {self.page_no()}", align="R")

# 5. Renderização do Painel Web (Identação Alinhada)
l_col1, l_col2 = st.columns([6, 4])

with l_col1:
    st.subheader("📋 Painel de Governança e Prazos")
    df_ed = df.copy().fillna("")
    df_ed.insert(0, "Excluir", False)
    
    cfg_cols = {
        "Excluir": st.column_config.CheckboxColumn(required=True),
        "Tarefa": st.column_config.TextColumn(width="large", default=""),
        "Início": st.column_config.TextColumn(alignment="center", default=""),
        "Conclusão": st.column_config.TextColumn(alignment="center", default=""),
        "Aprovador (A)": st.column_config.TextColumn(default=""),
        "Responsável (R)": st.column_config.TextColumn(default=""),
        "Consultados (C)": st.column_config.TextColumn(default=""),
        "Informados (I)": st.column_config.TextColumn(default="")
    }
    
    ed_res = st.data_editor(df_ed, use_container_width=True, hide_index=True, disabled=ordem, column_config=cfg_cols)
    
    if ed_res["Excluir"].any():
        manter = ed_res[~ed_res["Excluir"]]["Tarefa"].tolist()
        st.session_state.dados_estrategicos = st.session_state.dados_estrategicos[st.session_state.dados_estrategicos["Tarefa"].isin(manter)]
        st.rerun()

    if not df.empty:
        b1, b2 = st.columns(2)
        with b1:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as wr:
                df.to_excel(wr, index=False, sheet_name='Plano')
                ws = wr.sheets['Plano']
                f_c = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                fill_c = PatternFill('solid', fgColor='1B5E20')
                fill_z = PatternFill('solid', fgColor='F4FAF5')
                bd = Border(left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"), top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3"))
                
                for c in range(1, len(df.columns) + 1):
                    cell = ws.cell(1, c)
                    cell.font, cell.fill, cell.border = f_c, fill_c, bd
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                for r in range(2, len(df) + 2):
                    ws.row_dimensions[r].height = 20
                    for c in range(1, len(df.columns) + 1):
                        cell = ws.cell(r, c)
                        cell.font, cell.border = Font(name='Calibri', size=11), bd
                        cell.alignment = Alignment(horizontal='center' if c in [2,3,4,5,6,7] else 'left', vertical='center', wrap_text=True)
                        if r % 2 == 1: cell.fill = fill_z
                
                ws.conditional_formatting.add(f'E2:E{len(df)+1}', CellIsRule(operator='greaterThanOrEqual', formula=['60'], fill=PatternFill('solid', fgColor='FFEBEE'), font=Font(name='Calibri', size=11, bold=True, color='C62828')))
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 45 if col[0].column_letter == 'A' else max(max(len(str(cell.value or '')) for cell in col) + 4, 13)
                ws.row_dimensions[1].height = 26
            st.download_button("📄 Exportar Planilha Excel (.xlsx)", buf.getvalue(), "plano_trabalho_gut_raci.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            
        with b2:
            pdf = PDFExecutivo(orientation="L", unit="mm", format="A4")
            pdf.add_page()
            pdf.set_y(32)
            pdf.set_font("Arial", "B", 11)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 8, "I. Matriz de Priorização (GUT) e Distribuição de Papéis (RACI)", ln=True)
            pdf.ln(2)
            
            pdf.set_fill_color(46, 125, 50) 
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 8.5)
            
            w_tf, w_g, w_u, w_t, w_sc, w_dt, w_rc = 83, 7, 7, 7, 12, 22, 28
            pdf.set_x(12)
            pdf.cell(w_tf, 8, "Tarefa Estratégica", 1, 0, "", True)
            pdf.cell(w_g, 8, "G", 1, 0, "C", True); pdf.cell(w_u, 8, "U", 1, 0, "C", True); pdf.cell(w_t, 8, "T", 1, 0, "C", True)
            pdf.cell(w_sc, 8, "Score", 1, 0, "C", True)
            pdf.cell(w_dt, 8, "Início", 1, 0, "C", True); pdf.cell(w_dt, 8, "Conclusão", 1, 0, "C", True)
            pdf.cell(w_rc, 8, "Aprovador (A)", 1, 0, "", True); pdf.cell(w_rc, 8, "Responsável (R)", 1, 0, "", True)
            pdf.cell(w_rc, 8, "Consultado (C)", 1, 0, "", True); pdf.cell(w_rc, 8, "Informado (I)", 1, 1, "", True)
            
            for idx, row in df.iterrows():
                bg = idx % 2 == 1
                pdf.set_fill_color(244, 250, 245) if bg else pdf.set_fill_color(255, 255, 255)
                pdf.set_x(12)
                xi, yi = pdf.get_x(), pdf.get_y()
                pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 8)
                pdf.multi_cell(w_tf, 5, row['Tarefa'], 1, "L", bg)
                h_l = max(6, pdf.get_y() - yi)
                
                pdf.set_xy(xi + w_tf, yi)
                pdf.cell(w_g, h_l, str(row['G']), 1, 0, "C", bg); pdf.cell(w_u, h_l, str(row['U']), 1, 0, "C", bg); pdf.cell(w_t, h_l, str(row['T']), 1, 0, "C", bg)
                
                sc_v = int(row['Score'])
                if sc_v >= 60:
                    pdf.set_fill_color(255, 235, 238); pdf.set_text_color(198, 40, 40); pdf.set_font("Arial", "B", 8)
                pdf.cell(w_sc, h_l, str(sc_v), 1, 0, "C", True)
                
                pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "", 8)
                pdf.set_fill_color(244, 250, 245) if bg else pdf.set_fill_color(255, 255, 255)
                pdf.cell(w_dt, h_l, row['Início'], 1, 0, "C", bg); pdf.cell(w_dt, h_l, row['Conclusão'], 1, 0, "C", bg)
                pdf.cell(w_rc, h_l, str(row['Aprovador (A)']), 1, 0, "", bg); pdf.cell(w_rc, h_l, str(row['Responsável (R)']), 1, 0, "", bg)
                pdf.cell(w_rc, h_l, str(row['Consultados (C)']), 1, 0, "", bg); pdf.cell(w_rc, h_l, str(row['Informados (I)']), 1, 1, "", bg)
            
            # --- CRONOGRAMA EM CARDS (PÁGINA 2) ---
            pdf.add_page()
            pdf.set_y(32); pdf.set_font("Arial", "B", 11); pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 8, "II. Mapa de Prazos de Execução (Visão Linear)", ln=True)
            pdf.ln(3)
            
            for idx, row in df.iterrows():
                y_c_i = pdf.get_y()
                pdf.set_y(y_c_i + 2); pdf.set_x(18); pdf.set_font("Arial", "B", 9); pdf.set_text_color(0, 0, 0)
                pdf.multi_cell(255, 5, f"{idx+1}. {row['Tarefa']}", 0)
                pdf.ln(2)
                
                sc_v = int(row['Score'])
                pdf.set_fill_color(27, 94, 32) if sc_v >= 60 else (pdf.set_fill_color(44, 160, 90) if sc_v >= 30 else pdf.set_fill_color(161, 219, 178))
                w_b = max(15, min(100, sc_v * 1.3))
                pdf.rect(18, pdf.get_y() + 1, w_b, 4, "F")
                
                pdf.set_font("Arial", "", 8.5); pdf.set_text_color(80, 80, 80); pdf.set_x(23 + w_b)
                pdf.cell(0, 6, f"Prazo: {row['Início']} a {row['Conclusão']} | Responsável: {row['Responsável (R)']}", ln=True)
                
                y_c_f = pdf.get_y() + 2
                pdf.set_draw_color(200, 200, 200); pdf.set_line_width(0.2)
                pdf.rect(15, y_c_i, 267, y_c_f - y_c_i, "D")
                pdf.set_y(y_c_f + 4)
                
            pdf.set_draw_color(0, 0, 0)
            st.download_button("📕 Exportar Relatório PDF (.pdf)", bytes(pdf.output()), "relatorio_cronograma_tarefas.pdf", "application/pdf", use_container_width=True)

with l_col2:
    st.subheader("📊 Volume de Criticidade")
    if not df.empty:
        fig_bar = px.bar(df, x="Score", y="Tarefa", orientation="h", text="Score", color="Score", color_continuous_scale=["#a1dbb2", "#2ca05a", "#1b5e20"])
        fig_bar.update_yaxes(automargin=True)
        fig_bar.update_layout(yaxis={'categoryorder':'total ascending'}, showlegend=False, coloraxis_colorbar=dict(orientation="h", y=-0.3, title="GUT"), margin=dict(l=10, r=10, t=10, b=80), height=320)
        st.plotly_chart(fig_bar, use_container_width=True)
    else: st.info("Nenhuma tarefa registrada.")

# 6. Cronograma Interativo da Tela (Gantt)
st.markdown("---")
st.subheader("📅 Cronograma Interativo de Execução (Linha do Tempo)")
if not df_par_grafico.empty:
    df_tl = df_par_grafico.copy()
    df_tl["Início"] = pd.to_datetime(df_tl["Início"])
    df_tl["Conclusão"] = pd.to_datetime(df_tl["Conclusão"])
    
    f_gantt = px.timeline(df_tl, x_start="Início", x_end="Conclusão", y="Tarefa", color="Score", color_continuous_scale=["#a1dbb2", "#2ca05a", "#1b5e20"])
    f_gantt.update_xaxes(tickformat="%d/%m")
    f_gantt.update_yaxes(automargin=True)
    f_gantt.update_layout(yaxis={'categoryorder': 'total ascending'}, xaxis_title="Linha do Tempo (Dia/Mês)", yaxis_title="", coloraxis_colorbar=dict(orientation="h", y=-0.4), margin=dict(l=20, r=20, t=40, b=80), height=380)
    st.plotly_chart(f_gantt, use_container_width=True)
else: 
    st.info("Adicione tarefas para visualizar o mapa do cronograma.")
